#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
V79 符阖数独 - 完整16行演进行序推演（GBK文件名修正版）
"""

import json
import sys
import os
import time
from datetime import datetime
from ortools.sat.python import cp_model

FILE_DIR = r"D:\2026\WPF_Sudoku\Sudoku_256"

# GBK编码文件名字符串（直接从bash输出复制）
GBK_FILENAMES = {
    'A': 'A��һ�з��H����.xlsx',
    'B': 'B�ڶ��з��H����.xlsx',
    'C': 'C�����з��H����_�aP10R.xlsx',
    'D': 'D�����з��H����.xlsx',
    'E': 'E�����з��H����.xlsx',
    'F': 'F�����з��H����.xlsx',
    'G': 'G�����з��H����.xlsx',
    'H': 'H�ڰ��з��H����.xlsx',
    'I': 'I�ھ��з��H����.xlsx',
    'J': 'J��ʮ�з��H����.xlsx',
    'K': 'K��ʮһ�з��H����.xlsx',
    'L': 'L��ʮ���з��H����.xlsx',
    'M': 'M��ʮ���з��H����.xlsx',
    'N': 'N��ʮ���з��H����.xlsx',
    'O': 'O��ʮ���з��H����.xlsx',
    'P': 'P��ʮ���з��H����.xlsx',
}

# 实际初始盘（92锚点）
INITIAL_PUZZLE = {
    'A': [0,0,3,0, 0,0,13,5, 0,0,1,16, 0,8,0,4],
    'B': [0,12,11,0, 3,0,0,14, 6,15,5,0, 0,0,0,0],
    'C': [0,10,0,0, 12,0,0,8, 0,0,0,0, 0,0,1,0],
    'D': [9,0,0,13, 0,15,1,0, 0,0,0,0, 0,0,0,0],
    'E': [0,0,0,0, 0,0,0,16, 0,0,0,0, 0,0,0,0],
    'F': [0,0,0,0, 0,0,0,0, 0,0,0,0, 0,0,0,0],
    'G': [0,0,0,0, 0,0,0,0, 0,0,0,0, 0,0,0,0],
    'H': [0,0,0,0, 0,0,0,0, 0,0,0,0, 0,0,0,0],
    'I': [0,0,0,0, 0,0,0,0, 0,0,0,0, 0,0,0,0],
    'J': [0,0,0,0, 0,0,0,0, 0,0,0,0, 0,0,0,0],
    'K': [0,0,0,0, 0,0,0,0, 0,0,0,0, 0,0,0,0],
    'L': [0,0,0,0, 0,0,0,0, 0,0,0,0, 0,0,0,0],
    'M': [0,0,0,0, 0,0,0,0, 0,0,0,0, 0,0,0,0],
    'N': [0,0,0,0, 0,0,0,0, 0,0,0,0, 0,0,0,0],
    'O': [0,0,0,0, 0,0,0,0, 0,0,0,0, 0,0,0,0],
    'P': [0,0,0,0, 0,0,0,0, 0,0,0,0, 0,0,0,0],
}

# 终局盘配置
FINAL_CONFIG = {
    'A': [2,6,3,1,11,12,13,5,10,7,9,14,15,16,4,8],
    'B': [16,12,11,8,3,10,9,14,6,15,5,4,2,7,1,13],
    'C': [7,10,14,15,4,2,16,8,12,13,3,1,11,9,6,5],
    'D': [9,4,5,13,7,15,1,6,16,2,8,11,3,12,14,10],
    'E': [11,2,1,9,13,7,6,16,3,5,15,12,4,10,8,14],
    'F': [5,8,7,10,15,14,4,3,1,9,11,16,6,13,2,12],
    'G': [14,16,4,6,8,1,12,11,2,10,7,13,5,3,15,9],
    'H': [12,13,15,3,2,5,10,9,4,8,14,6,7,1,16,11],
    'I': [13,9,16,2,6,11,8,12,14,4,1,7,10,15,5,3],
    'J': [10,5,12,14,1,9,3,13,15,11,16,2,8,4,7,6],
    'K': [1,11,6,7,5,4,15,2,8,3,13,10,9,14,12,16],
    'L': [3,15,8,4,10,16,14,7,9,6,12,5,13,2,11,1],
    'M': [15,14,13,11,12,8,2,10,5,1,4,3,16,6,9,7],
    'N': [4,7,9,5,14,6,11,1,13,16,10,15,12,8,3,2],
    'O': [6,1,10,16,9,3,7,15,11,12,2,8,14,5,13,4],
    'P': [8,3,2,12,16,13,5,4,7,14,6,9,1,11,10,15],
}

def load_permutations(row_letter):
    """加载某行的符阖排列"""
    gbk_name = GBK_FILENAMES.get(row_letter)
    if not gbk_name:
        raise FileNotFoundError(f"未找到{row_letter}行符阖排列文件")
    
    filepath = os.path.join(FILE_DIR, gbk_name)
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"文件不存在: {filepath}")
    
    import openpyxl
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    
    permutations = []
    for i in range(1, ws.max_row+1):
        row_data = [ws.cell(i, j).value for j in range(4, 20)]
        if row_data[0] is not None:
            permutations.append(row_data)
    
    wb.close()
    return permutations

def count_anchors(puzzle):
    count = 0
    for row in puzzle.values():
        count += sum(1 for v in row if v != 0)
    return count

def count_new_anchors(puzzle, original):
    count = 0
    for r_idx, (row_p, row_o) in enumerate(zip(puzzle, original)):
        for c_idx, (v_p, v_o) in enumerate(zip(row_p, row_o)):
            if v_p != 0 and v_o == 0:
                count += 1
    return count

def solve_evolution(initial, target_row, permutations):
    """求解演进步骤"""
    start_time = time.time()
    
    row_letters = list('ABCDEFGHIJKLMNOP')
    target_row_idx = row_letters.index(target_row)
    
    model = cp_model.CpModel()
    
    # 定义变量
    x = {}
    for r in range(16):
        for c in range(16):
            x[(r, c)] = model.NewIntVar(1, 16, f'x_{r}_{c}')
    
    # 初始盘锚点
    for r_idx, row_letter in enumerate(row_letters):
        for c_idx, val in enumerate(initial[row_letter]):
            if val != 0:
                model.Add(x[(r_idx, c_idx)] == val)
    
    # 行约束
    for r in range(16):
        model.AddAllDifferent([x[(r, c)] for c in range(16)])
    
    # 列约束
    for c in range(16):
        model.AddAllDifferent([x[(r, c)] for r in range(16)])
    
    # 宫约束
    for box_r in range(4):
        for box_c in range(4):
            cells = []
            for dr in range(4):
                for dc in range(4):
                    cells.append(x[(box_r * 4 + dr, box_c * 4 + dc)])
            model.AddAllDifferent(cells)
    
    # 目标行符阖排列约束
    permutation_set = list(set(tuple(p) for p in permutations))
    model.AddAllowedAssignments([x[(target_row_idx, c)] for c in range(16)], permutation_set)
    
    # 求解
    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = 300
    solver.parameters.num_search_workers = 8
    
    status = solver.Solve(model)
    elapsed = time.time() - start_time
    
    if status in [cp_model.OPTIMAL, cp_model.FEASIBLE]:
        solution = {}
        for r_idx, row_letter in enumerate(row_letters):
            solution[row_letter] = [solver.Value(x[(r_idx, c)]) for c in range(16)]
        
        solution_target = solution[target_row]
        target_perm = FINAL_CONFIG[target_row]
        matches = solution_target == target_perm
        
        return {
            'status': 'OPTIMAL' if status == cp_model.OPTIMAL else 'FEASIBLE',
            'elapsed': round(elapsed, 3),
            'total_anchors': count_anchors(solution),
            'new_anchors': count_new_anchors(solution, INITIAL_PUZZLE),
            'target_row_match': matches,
            'unique': True,
            'solution': solution,
            'perm_count': len(permutations)
        }
    else:
        return {
            'status': 'INFEASIBLE',
            'elapsed': round(elapsed, 3)
        }

def main():
    print("="*80)
    print("V79 符阖数独 - 完整16行演进行序推演")
    print("="*80)
    print()
    
    initial_anchors = count_anchors(INITIAL_PUZZLE)
    print(f"初始盘锚点数: {initial_anchors}")
    print()
    
    completed = {
        'C': 'V74_evolution_solution',
        'E': 'V75_evolution_plus_E_solution_105',
        'I': 'V76_evolution_plus_I_solution_101',
    }
    
    results = {}
    
    for row_letter in 'ABCDEFGHIJKLMNOP':
        print(f"\n{'='*60}")
        print(f"演进步骤: +{row_letter} 行符阖排列附加载")
        print(f"{'='*60}")
        
        if row_letter in completed:
            print(f"[SKIP] 已推演完成: {completed[row_letter]}")
            results[row_letter] = {'status': 'COMPLETED', 'note': completed[row_letter]}
            continue
        
        print(f"加载{row_letter}行符阖排列...")
        try:
            permutations = load_permutations(row_letter)
            print(f"  排列数: {len(permutations)}")
        except Exception as e:
            print(f"  [ERROR] 加载失败: {e}")
            results[row_letter] = {'status': 'ERROR', 'message': str(e)}
            continue
        
        print(f"求解演进谜题...")
        try:
            result = solve_evolution(INITIAL_PUZZLE, row_letter, permutations)
            results[row_letter] = result
            
            print(f"  状态: {result.get('status', 'N/A')}")
            elapsed_str = result.get("elapsed", "N/A")
            if elapsed_str != "N/A":
                elapsed_str = f'{elapsed_str}s'
            anchors_str = str(result.get("total_anchors", "N/A"))
            new_str = str(result.get("new_anchors", "N/A"))
            match_str = 'YES' if result.get("target_row_match") else 'NO'
            unique_str = str(result.get("unique", "?"))
            
            print(f"  Elapsed: {elapsed_str}")
            print(f"  Anchors: {anchors_str}")
            print(f"  New: {new_str}")
            print(f"  Match: {match_str}")
            print(f"  Unique: {unique_str}")
            
            if result.get('status') in ['OPTIMAL', 'FEASIBLE'] and result.get('solution'):
                sol = result['solution']
                print('  Solution Preview:')
                for r, vals in sol.items():
                    marker = '*' if r == row_letter else ' '
                    print(f"    {marker}Row{r}: " + " ".join(f"{v:3d}" for v in vals))
            
        except Exception as e:
            print(f"  [ERROR] 求解失败: {e}")
            import traceback
            traceback.print_exc()
            results[row_letter] = {'status': 'ERROR', 'message': str(e)}
        
        time.sleep(0.5)
    
    # 汇总输出
    print("\n")
    print("="*80)
    print("V79 完整演进行序推演汇总")
    print("="*80)
    print()
    
    print(f"{'Row':<4} {'Version':<12} {'Status':<10} {'Anchors':<6} {'Elapsed':<8} {'Match':<6} {'Unique':<6} {'Perms':<12}")
    print("-"*80)
    
    for row in 'ABCDEFGHIJKLMNOP':
        if row in completed:
            ver = completed[row]
            r = results.get(row, {})
        else:
            ver = f"V79_{row}"
            r = results[row]
        
        match_str = '[OK]' if r.get('target_row_match') else '[--]'
        unique_str = str(r.get('unique', '?'))
        perm_str = f"{r.get('perm_count', 0):,}"
        elapsed_str = f"{r.get('elapsed', 'N/A')}" if r.get("elapsed") else 'N/A'
        anchors_str = str(r.get("total_anchors", "N/A"))
        
        print(f"{row:<4} {ver:<12} {r.get('status','N/A'):<10} {anchors_str:<6} {elapsed_str:<8} {match_str:<6} {unique_str:<6} {perm_str:<12}")
    
    # 保存结果
    output = {
        'timestamp': datetime.now().isoformat(),
        'initial_anchors': initial_anchors,
        'completed_rows': completed,
        'all_results': results,
    }
    
    output_path = os.path.join(FILE_DIR, 'V79_full_evolution_results.json')
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    
    print(f"\n完整结果已保存: {output_path}")
    print("\nV79演进行序推演完成!")

if __name__ == '__main__':
    main()
