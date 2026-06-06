#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""讀取關鍵Excel和txt檔案"""

import openpyxl
import json
from pathlib import Path

print("=" * 80)
print("讀取關鍵資料檔案")
print("=" * 80)

# 1. 讀取第16行符闔排列
print("\n" + "=" * 60)
print("1. P第十六行符闔排列.xlsx")
print("=" * 60)

wb = openpyxl.load_workbook('P第十六行符闔排列.xlsx', read_only=True)
ws = wb.active
print(f'Sheet名稱: {ws.title}')
print(f'最大列數: {ws.max_row}')
print(f'最大欄數: {ws.max_column}')

# 讀取所有資料
all_rows = []
for row in ws.iter_rows(values_only=True):
    all_rows.append(row)

print(f'\n總行數: {len(all_rows)}')
print('\n前5行資料:')
for i, row in enumerate(all_rows[:5]):
    print(f'  Row {i+1}: {row}')

print('\n最後3行資料:')
for i, row in enumerate(all_rows[-3:], len(all_rows)-2):
    print(f'  Row {i}: {row}')

wb.close()

# 2. 讀取遺傳傳遞概率
print("\n" + "=" * 60)
print("2. 遺傳傳遞概率.xlsx")
print("=" * 60)

wb2 = openpyxl.load_workbook('遺傳傳遞概率.xlsx', read_only=True)
ws2 = wb2.active
print(f'Sheet名稱: {ws2.title}')
print(f'最大列數: {ws2.max_row}')
print(f'最大欄數: {ws2.max_column}')

all_rows2 = []
for row in ws2.iter_rows(values_only=True):
    all_rows2.append(row)

print(f'\n總行數: {len(all_rows2)}')
print('\n所有資料:')
for i, row in enumerate(all_rows2):
    print(f'  Row {i+1}: {row}')

wb2.close()

# 3. 讀取超級大數獨_box_size4.txt
print("\n" + "=" * 60)
print("3. 超級大數獨_box_size4.txt")
print("=" * 60)

with open('超級大數獨_box_size4.txt', 'r', encoding='utf-8') as f:
    content = f.read()

print(f'檔案大小: {len(content)} 字節')
print('\n檔案內容預覽（前1000字元）:')
print(content[:1000])

# 儲存匯總資訊
print("\n" + "=" * 80)
print("匯總資訊")
print("=" * 80)
summary = {
    "P第十六行符闔排列": {
        "total_permutations": len(all_rows),
        "columns": len(all_rows[0]) if all_rows else 0,
        "description": f"第16行符闔排列，共{len(all_rows)}個排列"
    },
    "遺傳傳遞概率": {
        "rows": len(all_rows2),
        "columns": len(all_rows2[0]) if all_rows2 else 0,
        "description": "遺傳演算法中的傳遞概率資料"
    },
    "超級大數獨_box_size4": {
        "file_size_bytes": len(content),
        "description": "box_size=4的超級大數獨配置資料"
    }
}

with open('資料檔案匯總.json', 'w', encoding='utf-8') as f:
    json.dump(summary, f, indent=2, ensure_ascii=False)

print("\n匯總已保存到: 資料檔案匯總.json")
