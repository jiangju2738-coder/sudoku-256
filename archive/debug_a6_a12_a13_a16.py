#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
深度檢查 A6, A12, A13, A16 的資料結構
"""

from openpyxl import load_workbook

base_dir = "D:/2026/WPF_Sudoku/Sudoku_256"
remaining_files = {
    6: "A6第六行符闔排列.xlsx",
    12: "A12第十二行符闔排列.xlsx",
    13: "A13第十三行符闔排列.xlsx",
    16: "A16第十六行符闔排列.xlsx"
}

for row_idx, filename in remaining_files.items():
    filepath = f"{base_dir}/{filename}"
    
    print(f"\n{'='*70}")
    print(f"🔍 深度檢查第{row_idx}行: {filename}")
    print(f"{'='*70}")
    
    wb = load_workbook(filepath, read_only=True)
    ws = wb.active
    
    print(f"📐 工作表: {ws.title}, 最大欄位: {ws.max_column}")
    
    # 讀取前 10 行，找出資料模式
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= 10:
            break
        
        # 找數值在 1-16 範圍的欄位
        data_cols = []
        for j, val in enumerate(row):
            if isinstance(val, (int, float)) and 1 <= val <= 16:
                data_cols.append(j)
        
        print(f"\n  行{i}: 欄位數={len(row)}, 1-16數值在欄位: {data_cols[:20]}")
        
        if data_cols and len(data_cols) >= 15:
            print(f"        數值: {[row[c] for c in data_cols[:16]]}")
    
    wb.close()
