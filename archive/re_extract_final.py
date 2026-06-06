#!/usr/bin/env python3
"""重新提取符阖排列 - 最终修正版"""
import openpyxl, os, json, time

WORK_DIR = r"D:\2026\WPF_Sudoku\Sudoku_256"
os.chdir(WORK_DIR)

INITIAL = [
    [0,0,3,0,0,12,0,5,0,0,0,14,0,16,0,8],
    [0,12,0,0,3,0,9,0,6,0,5,4,2,0,1,0],
    [0,0,14,0,0,2,0,8,0,0,0,0,0,0,0,0],
    [0,4,0,13,7,0,1,0,0,0,0,11,0,12,0,0],
    [0,0,0,0,13,0,0,0,0,5,0,0,4,0,0,0],
    [0,8,0,0,15,0,4,3,0,9,0,0,0,13,0,12],
    [14,0,4,6,0,0,12,0,2,0,0,0,0,3,0,0],
    [0,13,0,0,0,5,0,9,0,0,14,6,0,0,16,0],
    [13,0,0,2,0,11,0,0,14,0,0,7,0,15,0,3],
    [0,5,0,0,0,0,0,0,0,0,16,0,8,0,7,0],
    [1,0,6,0,5,0,0,2,0,3,0,0,9,0,0,0],
    [0,0,0,4,0,16,14,0,0,0,12,5,0,0,0,1],
    [15,0,0,0,12,0,0,0,5,1,0,3,0,6,0,7],
    [0,0,9,0,0,6,0,0,13,0,0,15,0,0,3,0],
    [0,1,0,0,9,0,0,15,0,0,2,8,0,5,0,0],
    [0,0,2,0,0,0,5,0,0,14,0,0,1,0,10,15]
]

def extract_and_verify(row_num):
    """从xlsx正确提取排列并验证"""
    # 找到xlsx文件
    files = [f for f in os.listdir('.') if f.startswith(f'A{row_num}') and '符' in f and f.endswith('.xlsx')]
    if not files:
        return None, "file_not_found"
    
    filepath = files[0]
    wb = openpyxl.load_workbook(filepath, read_only=False, data_only=True)
    ws = wb.active
    
    # B-Q列 = 列2到列17 (16个排列数字)
    perms = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=2, max_col=17, values_only=True):
        perm = list(row)
        # 验证是有效排列
        if len(perm) == 16 and all(isinstance(v, (int, float)) and 1 <= v <= 16 for v in perm):
            if len(set(perm)) == 16:
                perms.append(perm)
    
    wb.close()
    
    # 验证与初始盘匹配
    known = {i: v for i, v in enumerate(INITIAL[row_num-1]) if v != 0}
    matches = sum(1 for p in perms if all(p[pos] == val for pos, val in known.items()))
    
    return perms, "ok" if matches == len(perms) else f"partial_{matches}"

# 处理所有16行
total = 0
for row_num in range(1, 17):
    perms, status = extract_and_verify(row_num)
    if perms is None:
        print(f"A{row_num}: {status}")
        continue
    
    total += len(perms)
    
    # 保存
    json_path = f"A{row_num}_permutations.json"
    with open(json_path, 'w') as f:
        json.dump(perms, f)
    
    known_count = sum(1 for v in INITIAL[row_num-1] if v != 0)
    print(f"A{row_num:2d}: {len(perms):7,} perms, {known_count} known, status={status} [{os.path.getsize(json_path)/1024/1024:.1f}MB]")

print(f"\n总计: {total:,} 排列")
