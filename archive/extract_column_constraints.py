#!/usr/bin/env python3
"""
从16个Excel文件中提取列约束数据
每行代表一个列约束，从E-T列提取16个数字
"""

import json
import os
import openpyxl
from collections import defaultdict

BASE = "D:/2026/WPF_Sudoku/Sudoku_256/"

def extract_permutations(filepath, row_idx):
    """提取单行16列排列"""
    permutations = []
    
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active
    
    # 找到对应的行（根据Excel文件名）
    target_row = row_idx
    found = False
    
    for row in ws.iter_rows(min_row=1, min_col=5, max_col=20, values_only=True):
        if not found:
            # 检查是否是对应的行
            found = True
        
        # 提取E-T列 (索引4-19) = 16个数字
        values = list(row)
        numeric_values = []
        
        for v in values:
            if v is None:
                continue
            if isinstance(v, (int, float)) and 1 <= v <= 16:
                numeric_values.append(int(v))
            elif isinstance(v, str) and v.startswith('='):
                continue
        
        if len(numeric_values) == 16:
            # 完整16个数字
            if all(1 <= x <= 16 for x in numeric_values):
                permutations.append(tuple(numeric_values))
        elif len(numeric_values) == 15:
            # 15个数字 + 1个ArrayFormula, 补上缺失的数字
            present = set(numeric_values)
            missing = [x for x in range(1, 17) if x not in present]
            if len(missing) == 1:
                perm = list(numeric_values)
                perm.append(missing[0])
                permutations.append(tuple(perm))
    
    wb.close()
    return permutations


def extract_in_chunks(filepath, row_idx, chunk_size=50000):
    """分块提取，内存友好"""
    all_perms = []
    
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active
    
    found = False
    
    for row in ws.iter_rows(min_row=1, min_col=5, max_col=20, values_only=True):
        if not found:
            found = True
        
        values = list(row)
        numeric_values = []
        
        for v in values:
            if v is None:
                continue
            if isinstance(v, (int, float)) and 1 <= v <= 16:
                numeric_values.append(int(v))
            elif isinstance(v, str) and v.startswith('='):
                continue
        
        if len(numeric_values) == 16:
            if all(1 <= x <= 16 for x in numeric_values):
                all_perms.append(tuple(numeric_values))
        elif len(numeric_values) == 15:
            present = set(numeric_values)
            missing = [x for x in range(1, 17) if x not in present]
            if len(missing) == 1:
                perm = list(numeric_values)
                perm.append(missing[0])
                all_perms.append(tuple(perm))
        
        if len(all_perms) % chunk_size == 0 and len(all_perms) > 0:
            print(f"  已处理 {len(all_perms):,} 个排列...")
    
    wb.close()
    return all_perms


def find_excel_file(row_num):
    """找到对应的Excel文件"""
    for f in os.listdir(BASE):
        if f.startswith(f"A{row_num}") and f.endswith(".xlsx"):
            return os.path.join(BASE, f)
    return None


def main():
    # 列约束：每行代表一列，共16列
    # 列约束应该从同样的Excel文件中提取，但理解方式不同
    
    # 注意：根据之前分析，E-T列就是16列数独数据
    # 每行Excel代表一个满足行约束的排列
    # 列约束需要从同样的数据中提取垂直方向的信息
    
    # 重新理解：用户说"同样从16个Excel文件"
    # 但之前提取的是行约束（每行Excel的一个排列）
    # 列约束需要理解Excel文件的不同结构
    
    # 让我先检查一个Excel文件的结构
    print("检查Excel文件结构...")
    sample_file = BASE + "A1第一行符闔排列.xlsx"
    wb = openpyxl.load_workbook(sample_file, read_only=True)
    ws = wb.active
    
    print(f"工作表: {ws.title}")
    print("\n前5行数据（所有列）:")
    for i, row in enumerate(ws.iter_rows(max_row=5, values_only=True)):
        print(f"Row {i+1}: {list(row)[:30]}")
    
    wb.close()
    
    # 根据之前的分析，每个Excel文件：
    # A列: 固定值（1, 2, ..., 16）- 表示行号
    # E-T列: 16个数字，表示该行的一个排列
    
    # 对于列约束，我们需要理解：
    # 1. 如果每个Excel文件代表一个固定的行位置
    # 2. 那么列约束应该从每个排列的相同位置提取
    
    # 让我检查文件命名和工作表名称
    print("\n检查工作表结构...")
    for i in range(1, 5):
        filepath = find_excel_file(i)
        if filepath:
            wb = openpyxl.load_workbook(filepath, read_only=True)
            print(f"A{i}: 工作表名 = {wb.sheetnames}")
            ws = wb.active
            # 检查列结构
            for j, row in enumerate(ws.iter_rows(max_row=1, values_only=True)):
                print(f"  第1行: {[v for v in row[:25]]}")
                break
            wb.close()


if __name__ == "__main__":
    main()
