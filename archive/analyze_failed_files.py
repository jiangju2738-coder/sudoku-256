#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
深度分析 A4-A6, A12-A16 檔案結構問題
找出資料讀取失敗的原因
"""

import sys
sys.path.insert(0, '.')

from openpyxl import load_workbook
import pandas as pd
import numpy as np

base_dir = "D:/2026/WPF_Sudoku/Sudoku_256"

# 未成功載入的檔案列表
failed_files = {
    4: "A4第四行符闔排列.xlsx",
    5: "A5第五行符闔排列.xlsx",
    6: "A6第六行符闔排列.xlsx",
    12: "A12第十二行符闔排列.xlsx",
    13: "A13第十三行符闔排列.xlsx",
    14: "A14第十四行符闔排列.xlsx",
    15: "A15第十五行符闔排列.xlsx",
    16: "A16第十六行符闔排列.xlsx"
}

print("=" * 70)
print("🔍 深度分析未成功載入的 8 個檔案")
print("=" * 70)
print()

for row_idx, filename in failed_files.items():
    filepath = f"{base_dir}/{filename}"
    
    print(f"\n{'='*70}")
    print(f"📊 分析第{row_idx}行: {filename}")
    print(f"{'='*70}")
    
    try:
        wb = load_workbook(filepath, read_only=True)
        ws = wb.active
        
        print(f"  📐 工作表名: {ws.title}")
        print(f"  📏 最大行: {ws.max_row}, 最大列: {ws.max_column}")
        
        # 讀取前 5 行，詳細分析每列的資料類型
        rows_data = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            rows_data.append(row)
            if i >= 5:
                break
        
        print(f"\n  📋 前 5 行原始資料:")
        for i, row in enumerate(rows_data):
            print(f"    行{i}:")
            for j, val in enumerate(row):
                if val is not None:
                    val_type = type(val).__name__
                    if isinstance(val, str) and val.startswith('='):
                        print(f"      列{j}: \"{val}\" [{val_type}]")
                    elif isinstance(val, (int, float)):
                        print(f"      列{j}: {val} [{val_type}]")
                    else:
                        print(f"      列{j}: {str(val)[:20]} [{val_type}]")
                else:
                    print(f"      列{j}: None")
        
        # 分析是否有數值資料
        numeric_data = []
        for row in rows_data:
            for val in row:
                if isinstance(val, (int, float)) and 1 <= val <= 16:
                    numeric_data.append(val)
        
        if numeric_data:
            print(f"\n  📊 發現 1-16 範圍的數值: {len(numeric_data)} 個")
            print(f"     範例: {numeric_data[:20]}")
        else:
            print(f"\n  ⚠ 未發現 1-16 範圍的數值")
        
        # 檢查是否有其他可能的資料範圍
        all_numeric = [val for row in rows_data for val in row if isinstance(val, (int, float))]
        if all_numeric:
            print(f"  📊 所有數值範圍: {min(all_numeric)} - {max(all_numeric)}")
        
        # 檢查是否有公式引用的資料
        formula_cols = []
        for col_idx in range(len(rows_data[0])):
            for row in rows_data:
                val = row[col_idx]
                if isinstance(val, str) and val.startswith('='):
                    formula_cols.append(col_idx)
                    break
        
        if formula_cols:
            print(f"  🔢 公式欄位: {formula_cols[:10]}")
        
        wb.close()
        
    except Exception as e:
        print(f"  ✗ 讀取失敗: {type(e).__name__}: {e}")

print("\n" + "=" * 70)
print("✅ 分析完成")
print("=" * 70)
