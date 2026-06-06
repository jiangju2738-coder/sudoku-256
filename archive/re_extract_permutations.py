#!/usr/bin/env python3
"""
重新提取符阖排列 - 基于xlsx数据源（与初始盘一致）
修正: 从B-Q列(索引2-17)提取16个排列数字
"""

import openpyxl, os, json, time

WORK_DIR = r"D:\2026\WPF_Sudoku\Sudoku_256"
os.chdir(WORK_DIR)

INITIAL_PUZZLE = [
    [0,0,3,0,0,12,0,5,0,0,0,14,0,16,0,8],
    [0,12,0,0,3,0,9,0,6,0,5,4,2,0,1,0],
    [0,0,14,0,0,2,0,8,0,0,0,0,0,0,0,0],
    [0,4,0,13,7,0,1,0,0,0,0,11,0,12,0,0],
    [0,0,0,0,13,0,0,0,0,5,0,0,4,0,0,0],
    [0,8,0,0,15,0,4,3,0,9,0,0,0,13,0,12],
    [14,0,4,6,0,0,12,0,2,0,0,0,0,3,0,0],
    [0,13,0,0,0,5,0,9,0,0,14,6,0,0,16,0],
    [13,0,0,2,0,11,0,0,14,0,0,7,0,15,0,3],
    [0,5,0,0,0,0,0,0,0,0,16,0,8,0,7,0],
    [1,0,6,0,5,0,0,2,0,3,0,0,9,0,0,0],
    [0,0,0,4,0,16,14,0,0,0,12,5,0,0,0,1],
    [15,0,0,0,12,0,0,0,5,1,0,3,0,6,0,7],
    [0,0,9,0,0,6,0,0,13,0,0,15,0,0,3,0],
    [0,1,0,0,9,0,0,15,0,0,2,8,0,5,0,0],
    [0,0,2,0,0,0,5,0,0,14,0,0,1,0,10,15]
]

def extract_perms(filepath):
    wb = openpyxl.load_workbook(filepath, read_only=False, data_only=True)
    ws = wb.active
    perms = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=2, max_col=17, values_only=True):
        perm = [v for v in row if isinstance(v, (int, float))]
        if len(perm) == 16 and all(1 <= v <= 16 for v in perm) and len(set(perm)) == 16:
            perms.append(perm)
    wb.close()
    return perms

def main():
    print("="*70)
    print("符阖排列重新提取 (xlsx数据源)")
    print("="*70)
    
    all_perms = {}
    results = []
    
    for row_idx in range(16):
        row_num = row_idx + 1
        xlsx_files = [f for f in os.listdir('.') if f.startswith(f'A{row_num}') and '符' in f]
        if not xlsx_files:
            continue
        
        perms = extract_perms(xlsx_files[0])
        all_perms[row_num] = perms
        
        known = {i:v for i,v in enumerate(INITIAL_PUZZLE[row_idx]) if v != 0}
        matches = sum(1 for p in perms if all(p[pos]==val for pos,val in known.items()))
        
        print(f"A{row_num:2d}: {len(perms):7,} perms, {matches:,} matches {'OK' if matches==len(perms) else 'CONFLICT'}")
        
        with open(f'A{row_num}_permutations.json','w') as f:
            json.dump(perms, f)
        
        results.append({"row":row_num,"count":len(perms),"matches":matches})
    
    # 列约束
    col_sets = [set() for _ in range(16)]
    for perms in all_perms.values():
        for p in perms:
            for i,v in enumerate(p):
                col_sets[i].add(v)
    
    print("\n列约束分析:")
    for i in range(16):
        missing = [v for v in range(1,17) if v not in col_sets[i]]
        if missing:
            print(f"  列{i+1:2d}: 缺失 {missing}")
    
    # 初始盘列约束验证
    non_compliant = []
    for ri in range(16):
        for ci,v in enumerate(INITIAL_PUZZLE[ri]):
            if v and v not in col_sets[ci]:
                non_compliant.append((ri+1,ci+1,v))
    
    if non_compliant:
        print(f"\n列约束冲突: {len(non_compliant)}个")
        for r,c,v in non_compliant:
            print(f"  ({r},{c}): {v}")
    else:
        print("\n✅ 所有92个已知数字符合列约束!")
    
    total = sum(r["count"] for r in results)
    print(f"\n总计: {total:,} 排列")
    print("JSON文件已更新")

if __name__ == "__main__":
    main()
