#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
專門提取 A16 的排列資料
A16 的結構: 欄位 4-16 (13 個數值) + 欄位 18 (1 個數值) + 欄位 19 (2 個數值) = 16 個
"""

from openpyxl import load_workbook
import numpy as np

base_dir = "D:/2026/WPF_Sudoku/Sudoku_256"

def collect_permutation_a16(row):
    """A16: 欄位 4-16 + 欄位 18 + 欄位 19 提取"""
    values = []
    
    # 欄位 4-16 (13 個數值)
    for col in range(4, 17):
        if col < len(row):
            val = row[col]
            if isinstance(val, (int, float)) and 1 <= val <= 16:
                values.append(int(val))
    
    # 欄位 18 (1 個數值)
    if 18 < len(row):
        val = row[18]
        if isinstance(val, (int, float)) and 1 <= val <= 16:
            values.append(int(val))
    
    # 欄位 19 (1 個數值，欄位 20 可能超出)
    if 19 < len(row):
        val = row[19]
        if isinstance(val, (int, float)) and 1 <= val <= 16:
            values.append(int(val))
    
    # 如果只有 15 個，補齊
    if len(values) == 15:
        present = set(values)
        missing = [x for x in range(1, 17) if x not in present]
        if len(missing) == 1:
            values.append(missing[0])
    
    return values if len(values) == 16 else None


# 讀取 A16
filepath = f"{base_dir}/A16第十六行符闔排列.xlsx"
print("📊 專門提取 A16 第十六行")
print("=" * 60)

wb = load_workbook(filepath, read_only=True)
ws = wb.active

# 分析前 5 行
print("\n🔍 分析前 5 行的資料:")
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i >= 5:
        break
    
    print(f"\n  行{i}:")
    
    # 欄位 4-20
    for col in range(4, min(21, len(row))):
        val = row[col]
        if val is not None:
            if isinstance(val, (int, float)):
                print(f"    欄位{col:2d}: {int(val):3d} {'✓' if 1 <= val <= 16 else ''}")
            else:
                print(f"    欄位{col:2d}: {str(val)[:25]}")
        else:
            print(f"    欄位{col:2d}: None")

wb.close()

# 正式提取
wb = load_workbook(filepath, read_only=True)
ws = wb.active

print("\n🚀 提取排列...")
perms = []
total_rows = 0

for row in ws.iter_rows(values_only=True):
    total_rows += 1
    perm = collect_permutation_a16(row)
    if perm:
        perms.append(perm)
    
    if total_rows > 20000:
        break

wb.close()

print(f"\n{'='*60}")
if perms:
    print(f"✓ 成功讀取 {len(perms):,} 個排列模式")
    for i, perm in enumerate(perms[:5]):
        present = set(perm)
        missing = [x for x in range(1, 17) if x not in present]
        duplicates = [x for x in present if perm.count(x) > 1]
        status = "✓" if not missing and not duplicates else f"⚠ 缺失{missing}, 重複{duplicates}"
        print(f"  範例{i+1}: {perm} {status}")
else:
    print("⚠ 未找到有效排列")

# 儲存結果
import json
with open(f"{base_dir}/A16_提取結果.json", 'w') as f:
    json.dump({"count": len(perms), "samples": perms[:10]}, f, ensure_ascii=False, indent=2)

print(f"✓ 結果已儲存到 A16_提取結果.json")
