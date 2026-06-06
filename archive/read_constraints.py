#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
快速讀取 16 行符闔排列 Excel 檔案
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook

base_dir = "D:/2026/WPF_Sudoku/Sudoku_256"

# 嘗試用 openpyxl 快速讀取第一個檔案
print("嘗試讀取 A1 檔案...")
try:
    wb = load_workbook(f"{base_dir}/A1第一行符闔排列.xlsx", read_only=True)
    ws = wb.active
    print(f"工作表名: {ws.title}")
    print(f"最大行: {ws.max_row}, 最大列: {ws.max_column}")
    
    # 讀取前幾行
    rows_data = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        rows_data.append(row)
        if i >= 5:
            break
    
    print(f"\n前 {len(rows_data)} 行資料:")
    for i, row in enumerate(rows_data):
        print(f"  行{i}: {row[:10]}...")  # 只顯示前10列
    
    wb.close()
except Exception as e:
    print(f"openpyxl 讀取失敗: {e}")
    
    # 回退到 pandas
    try:
        df = pd.read_excel(f"{base_dir}/A1第一行符闔排列.xlsx", engine='openpyxl', nrows=5)
        print(f"\nPandas 讀取成功, shape: {df.shape}")
        print(f"欄位: {list(df.columns)[:10]}")
        print(f"\n前幾行:")
        print(df.head())
    except Exception as e2:
        print(f"Pandas 也失敗: {e2}")
