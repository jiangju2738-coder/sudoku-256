#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
專門分析 A6, A12, A13, A16 的欄位結構
找出真正包含排列資料的欄位
"""

from openpyxl import load_workbook
from collections import Counter

base_dir = "D:/2026/WPF_Sudoku/Sudoku_256"
target_files = {
    6: "A6第六行符闔排列.xlsx",
    12: "A12第十二行符闔排列.xlsx",
    13: "A13第十三行符闔排列.xlsx",
    16: "A16第十六行符闔排列.xlsx"
}

for row_idx, filename in target_files.items():
    filepath = f"{base_dir}/{filename}"
    
    print(f"\n{'='*70}")
    print(f"🔍 詳細分析第{row_idx}行: {filename}")
    print(f"{'='*70}")
    
    wb = load_workbook(filepath, read_only=True)
    ws = wb.active
    
    # 分析前 10 行的每個欄位
    all_rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        all_rows.append(row)
        if i >= 10:
            break
    
    wb.close()
    
    # 顯示前 10 行的所有欄位值
    print(f"\n📋 前 10 行完整資料:")
    for i, row in enumerate(all_rows):
        print(f"\n  行{i}:")
        
        # 分析每個欄位
        for j in range(min(25, len(row))):
            val = row[j]
            if val is not None:
                if isinstance(val, (int, float)):
                    print(f"    欄{j:2d}: {int(val):3d} ", end="")
                    if j >= 4 and j <= 19:
                        print(f"[欄{4-19}]", end="")
                    print()
                else:
                    print(f"    欄{j:2d}: \"{str(val)[:15]}\"", end="")
                    if isinstance(val, str) and val.startswith('='):
                        print(" [公式]", end="")
                    print()
            else:
                print(f"    欄{j:2d}: None")
    
    # 專門分析欄位 4-19 的資料
    print(f"\n🎯 欄位 4-19 分析:")
    for j in range(4, 20):
        if j < len(all_rows[0]):
            values = [row[j] for row in all_rows if j < len(row) and isinstance(row[j], (int, float)) and 1 <= row[j] <= 16]
            if values:
                unique_vals = set(values)
                print(f"  欄{j}: {len(unique_vals)} 個唯一值, 範例: {list(unique_vals)[:8]}")
            else:
                print(f"  欄{j}: 無有效資料")
        else:
            print(f"  欄{j}: 超出範圍")
    
    # 查找真正的排列欄位
    print(f"\n🔎 尋找排列模式:")
    for i, row in enumerate(all_rows[:5]):
        # 查找欄位 4-19 中 1-16 的唯一值
        if len(row) >= 20:
            row_values = []
            for j in range(4, 20):
                val = row[j]
                if isinstance(val, (int, float)) and 1 <= val <= 16:
                    row_values.append(int(val))
                elif isinstance(val, str) and val.startswith('='):
                    break
            
            if len(row_values) == 16:
                present = set(row_values)
                missing = [x for x in range(1, 17) if x not in present]
                duplicates = [x for x in present if row_values.count(x) > 1]
                print(f"  行{i}: 欄位4-19有{len(row_values)}個值")
                print(f"        數值: {row_values}")
                print(f"        唯一值: {len(present)}, 重複: {duplicates}, 缺失: {missing}")
            
            # 檢查其他可能的欄位位置
            all_numeric_in_row = [val for val in row if isinstance(val, (int, float)) and 1 <= val <= 16]
            if len(all_numeric_in_row) >= 16:
                unique_all = set(all_numeric_in_row)
                print(f"  行{i}: 整行有{len(all_numeric_in_row)}個1-16數值, 唯一值{len(unique_all)}個")
