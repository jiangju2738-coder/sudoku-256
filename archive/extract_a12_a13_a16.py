#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
專門提取 A12, A13, A16 的排列資料
這些檔案的欄位 18 是 ArrayFormula，資料分佈在欄位 4-17 + 欄位 18 + 欄位 19
"""

from openpyxl import load_workbook
import numpy as np

base_dir = "D:/2026/WPF_Sudoku/Sudoku_256"

def collect_permutation_from_cols(row, row_idx):
    """
    從特定欄位組合收集排列
    A12/A13: 欄位 4-17 (14 個) + 欄位 18 (公式結果) + 欄位 19 (1 個) = 16 個
    A16: 欄位 4-16 (13 個) + 欄位 18 (公式結果) + 欄位 19 (2 個) = 16 個
    """
    values = []
    
    if row_idx in [12, 13]:
        # A12/A13: 欄位 4-17 有 14 個數值，欄位 18 是 ArrayFormula，欄位 19 有 1 個數值
        # 欄位 4-17
        for col in range(4, 18):
            if col < len(row):
                val = row[col]
                if isinstance(val, (int, float)) and 1 <= val <= 16:
                    values.append(int(val))
        
        # 欄位 18 可能是公式，但欄位 19 有數值
        if 19 < len(row):
            val = row[19]
            if isinstance(val, (int, float)) and 1 <= val <= 16:
                values.append(int(val))
        
        # 如果只有 15 個，補齊
        if len(values) == 15:
            present = set(values)
            missing = [x for x in range(1, 17) if x not in present]
            if len(missing) == 1:
                values.append(missing[0])
    
    elif row_idx == 16:
        # A16: 欄位 4-16 (13 個數值) + 欄位 18 (公式結果) + 欄位 19 (1 個數值)
        # 欄位 4-16
        for col in range(4, 17):
            if col < len(row):
                val = row[col]
                if isinstance(val, (int, float)) and 1 <= val <= 16:
                    values.append(int(val))
        
        # 欄位 18 是公式
        # 欄位 19-20 有數值
        for col in [19, 20]:
            if col < len(row):
                val = row[col]
                if isinstance(val, (int, float)) and 1 <= val <= 16:
                    values.append(int(val))
        
        # 如果只有 15 個，補齊
        if len(values) == 15:
            present = set(values)
            missing = [x for x in range(1, 17) if x not in present]
            if len(missing) == 1:
                values.append(missing[0])
    
    return values if len(values) == 16 else None


def analyze_specific_row(filepath, row_idx, num_rows=5):
    """分析特定行的詳細欄位結構"""
    wb = load_workbook(filepath, read_only=True)
    ws = wb.active
    
    print(f"\n🔍 分析第{row_idx}行前 {num_rows} 行的欄位結構:")
    
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        print(f"\n  行{i}:")
        
        # 分析欄位 4-25
        for col in range(4, min(26, len(row))):
            val = row[col]
            if val is not None:
                if isinstance(val, (int, float)):
                    marker = "✓" if 1 <= val <= 16 else ""
                    print(f"    欄位{col:2d}: {int(val):3d} {marker}")
                else:
                    print(f"    欄位{col:2d}: \"{str(val)[:25]}\"")
            else:
                print(f"    欄位{col:2d}: None")
        
        if i >= num_rows - 1:
            break
    
    wb.close()


def load_specific_file(filepath, row_idx):
    """載入特定檔案"""
    wb = load_workbook(filepath, read_only=True)
    ws = wb.active
    
    row_permutations = []
    total_rows = 0
    
    for row in ws.iter_rows(values_only=True):
        total_rows += 1
        perm = collect_permutation_from_cols(row, row_idx)
        if perm:
            row_permutations.append(perm)
        
        if total_rows > 30000:
            break
    
    wb.close()
    return row_permutations


if __name__ == "__main__":
    print("=" * 70)
    print("🎯 專門提取 A12, A13, A16")
    print("=" * 70)
    
    results = {}
    
    for row_idx in [12, 13, 16]:
        row_names = {12: "第十二", 13: "第十三", 16: "第十六"}
        filename = f"A{row_idx}{row_names[row_idx]}行符闔排列.xlsx"
        filepath = f"{base_dir}/{filename}"
        
        print(f"\n{'='*70}")
        print(f"📂 第{row_idx}行: {filename}")
        print(f"{'='*70}")
        
        # 先分析結構
        analyze_specific_row(filepath, row_idx, 3)
        
        # 載入資料
        print(f"\n🚀 提取排列...")
        perms = load_specific_file(filepath, row_idx)
        
        if perms:
            results[row_idx] = np.array(perms)
            print(f"✓ 成功讀取 {len(perms):,} 個排列模式")
            for i, perm in enumerate(perms[:3]):
                present = set(perm)
                missing = [x for x in range(1, 17) if x not in present]
                duplicates = [x for x in present if perm.count(x) > 1]
                status = "✓" if not missing and not duplicates else f"⚠ 缺失{missing}, 重複{duplicates}"
                print(f"  範例{i+1}: {perm} {status}")
        else:
            print(f"⚠ 未找到有效排列")
    
    # 總結
    print("\n" + "=" * 70)
    print("📊 結果彙總")
    print("=" * 70)
    
    total = 0
    for row_idx in [12, 13, 16]:
        if row_idx in results:
            count = len(results[row_idx])
            total += count
            print(f"   ✓ 第{row_idx:2d}行: {count:>8,} 個排列")
        else:
            print(f"   ⚠ 第{row_idx:2d}行: 0 個排列")
    
    print(f"\n📈 總計: {total:,} 個排列")
    
    # 儲存
    import json
    summary = {str(k): len(v) for k, v in results.items()}
    with open(f"{base_dir}/A12_A13_A16_提取結果.json", 'w') as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)
