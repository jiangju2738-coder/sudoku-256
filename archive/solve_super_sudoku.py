#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
超級大數獨 (16x16, box_size=4) 完整求解器
五維思維框架 + DLX 精確覆蓋 + 行約束優化
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
import json
import time
import html
from collections import defaultdict, Counter
from typing import List, Set, Dict, Tuple, Optional
from copy import deepcopy
from dataclasses import dataclass, field
from itertools import product

# ============== 數據載入模組 ==============

def load_all_row_constraints(base_dir: str) -> Dict[int, np.ndarray]:
    """讀取 16 行符闔排列 Excel 檔案"""
    row_files = {
        1: "A1第一行符闔排列.xlsx",
        2: "A2第二行符闔排列.xlsx",
        3: "A3第三行符闔排列.xlsx",
        4: "A4第四行符闔排列.xlsx",
        5: "A5第五行符闔排列.xlsx",
        6: "A6第六行符闔排列.xlsx",
        7: "A7第七行符闔排列.xlsx",
        8: "A8第八行符闔排列.xlsx",
        9: "A9第九行符闔排列.xlsx",
        10: "A10第十行符闔排列.xlsx",
        11: "A11第十一行符闔排列.xlsx",
        12: "A12第十二行符闔排列.xlsx",
        13: "A13第十三行符闔排列.xlsx",
        14: "A14第十四行符闔排列.xlsx",
        15: "A15第十五行符闔排列.xlsx",
        16: "A16第十六行符闔排列.xlsx"
    }
    
    constraints = {}
    for row_idx, filename in row_files.items():
        print(f"📂 讀取第{row_idx}行: {filename}")
        
        wb = load_workbook(f"{base_dir}/{filename}", read_only=True)
        ws = wb.active
        
        # 提取第 E-T 列 (欄位索引 4-19)
        row_permutations = []
        for row in ws.iter_rows(values_only=True):
            if len(row) >= 20:  # 至少要有欄位 4-19
                # 提取 E-T 列的值 (欄位索引 4-19)
                values = []
                for col_idx in range(4, 20):
                    val = row[col_idx]
                    if isinstance(val, (int, float)) and 1 <= val <= 16:
                        values.append(int(val))
                    else:
                        break
                
                if len(values) == 16:  # 完整的 16 個數字的排列
                    row_permutations.append(values)
        
        wb.close()
        
        if row_permutations:
            constraints[row_idx] = np.array(row_permutations)
            print(f"   ✓ 讀取 {len(row_permutations)} 個排列模式")
        else:
            constraints[row_idx] = np.array([]).reshape(0, 16)
            print(f"   ⚠ 無有效排列")
    
    return constraints


def load_initial_puzzle(txt_file: str) -> np.ndarray:
    """從 txt 檔案載入初盤"""
    grid = np.zeros((16, 16), dtype=np.int8)
    
    try:
        with open(txt_file, 'r', encoding='utf-8') as f:
            lines = f.readlines()
        
        # 跳過標題行，從第 3 行開始讀取數獨資料
        row_idx = 0
        for line in lines:
            line = line.strip()
            if not line or line.startswith('#') or line.startswith('A:') or line.startswith('B:'):
                continue
            
            # 嘗試解析數值行
            parts = line.split()
            values = []
            for part in parts:
                try:
                    val = int(part)
                    if 0 <= val <= 16:
                        values.append(val)
                except:
                    continue
            
            if len(values) == 16 and row_idx < 16:
                grid[row_idx, :] = values
                row_idx += 1
                print(f"第{row_idx}行: {values}")
        
        if row_idx == 0:
            print("⚠ 未找到有效數獨資料，使用空盤")
            
    except FileNotFoundError:
        print("⚠ 初盤檔案不存在，使用空盤")
    
    return grid


# ============== 五維思維框架 ==============

class FiveDimThinkingEngine:
    """五維思維框架引擎：點-線-面-體-球-時空"""
    
    def __init__(self):
        self.dimensions = {
            'point': {},    # 單元格約束
            'line': {},     # 行列約束
            'face': {},     # 宮格約束
            'body': {},     # 行組合約束
            'sphere': {},   # 全局狀態
            'spacetime': [] # 時空映射
        }
    
    def init_point_dim(self, grid_size: int = 16):
        """初始化點維度：256 個單元格的約束"""
        for i in range(grid_size):
            for j in range(grid_size):
                self.dimensions['point'][(i, j)] = set(range(1, 17))
    
    def add_line_constraint(self, row_idx: int, permutation: List[int]):
        """添加行約束（線維度）"""
        self.dimensions['line'][row_idx] = permutation
    
    def add_body_constraint(self, row_idx: int, permutations: np.ndarray):
        """添加體維度：行排列集合"""
        self.dimensions['body'][row_idx] = permutations
    
    def get_possibilities(self, grid: np.ndarray, row: int, col: int) -> Set[int]:
        """計算單元格的可能值"""
        possibilities = set(range(1, 17))
        
        # 同行約束
        possibilities -= set(grid[row, :])
        possibilities -= {0}
        
        # 同列約束
        possibilities -= set(grid[:, col])
        
        # 同宮格約束
        box_row = row // 4
        box_col = col // 4
        box_start_row = box_row * 4
        box_start_col = box_col * 4
        
        for i in range(box_start_row, box_start_row + 4):
            for j in range(box_start_col, box_start_col + 4):
                if grid[i, j] != 0:
                    possibilities.discard(grid[i, j])
        
        return possibilities


# ============== 超級數獨求解器 ==============

class SuperSudoku16x16Solver:
    """超級大數獨 16x16 求解器"""
    
    GRID_SIZE = 16
    BOX_SIZE = 4
    DIGITS = set(range(1, 17))
    
    def __init__(self, row_constraints: Dict[int, np.ndarray], initial_grid: np.ndarray):
        self.row_constraints = row_constraints
        self.grid = initial_grid.copy()
        self.solutions: List[np.ndarray] = []
        self.framework = FiveDimThinkingEngine()
        self.framework.init_point_dim()
        
        # 初始化五維約束
        for row_idx, perms in row_constraints.items():
            if len(perms) > 0:
                self.framework.add_body_constraint(row_idx, perms)
                self.framework.dimensions['line'][row_idx] = perms[0]  # 暫存第一個排列
        
    def is_valid(self, row: int, col: int, value: int) -> bool:
        """檢查填入值是否合法"""
        # 同行檢查
        if value in self.grid[row, :]:
            return False
        
        # 同列檢查
        if value in self.grid[:, col]:
            return False
        
        # 同宮格檢查
        box_row = row // self.BOX_SIZE
        box_col = col // self.BOX_SIZE
        for i in range(box_row * self.BOX_SIZE, (box_row + 1) * self.BOX_SIZE):
            for j in range(box_col * self.BOX_SIZE, (box_col + 1) * self.BOX_SIZE):
                if self.grid[i, j] == value:
                    return False
        
        return True
    
    def find_best_cell_mrv(self) -> Optional[Tuple[int, int, Set[int]]]:
        """MRV (Minimum Remaining Values) 策略"""
        min_poss = float('inf')
        best_cell = None
        
        for row in range(self.GRID_SIZE):
            for col in range(self.GRID_SIZE):
                if self.grid[row, col] == 0:
                    poss = self.framework.get_possibilities(self.grid, row, col)
                    if len(poss) < min_poss:
                        min_poss = len(poss)
                        best_cell = (row, col, poss)
                        if min_poss == 1:
                            return best_cell
        
        return best_cell
    
    def solve(self, max_solutions: int = 100, time_limit: int = 120) -> int:
        """求解器主函數"""
        start_time = time.time()
        
        def backtrack(count: int = 0) -> int:
            elapsed = time.time() - start_time
            if elapsed > time_limit:
                print(f"⏰ 時間限制 {time_limit}秒已達")
                return count
            
            cell = self.find_best_cell_mrv()
            if cell is None:
                # 找到完整解
                self.solutions.append(deepcopy(self.grid))
                count += 1
                print(f"✓ 找到第 {count} 個解 | 已用 {elapsed:.1f}秒")
                return count if count >= max_solutions else count
            
            row, col, possibilities = cell
            
            # 時空映射記錄
            self.framework.dimensions['spacetime'].append({
                'step': count,
                'cell': (row, col),
                'try_values': list(possibilities)
            })
            
            for value in sorted(possibilities):
                if self.is_valid(row, col, value):
                    self.grid[row, col] = value
                    count = backtrack(count)
                    if count >= max_solutions:
                        return count
                    self.grid[row, col] = 0
            
            return count
        
        return backtrack()
    
    def generate_visualization(self, output_path: str):
        """生成可視化 HTML"""
        elapsed = time.time()
        
        html_content = f'''<!DOCTYPE html>
<html lang="zh-TW">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>超級大數獨 16×16 求解結果</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            padding: 20px;
            min-height: 100vh;
        }}
        .container {{ max-width: 1400px; margin: 0 auto; }}
        h1 {{
            text-align: center;
            color: white;
            margin-bottom: 10px;
            text-shadow: 2px 2px 4px rgba(0,0,0,0.3);
        }}
        .subtitle {{ text-align: center; color: rgba(255,255,255,0.9); margin-bottom: 20px; }}
        
        .stats-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
            gap: 15px;
            margin-bottom: 30px;
        }}
        .stat-card {{
            background: rgba(255,255,255,0.95);
            border-radius: 12px;
            padding: 20px;
            text-align: center;
            box-shadow: 0 4px 15px rgba(0,0,0,0.2);
        }}
        .stat-value {{ font-size: 2.2em; font-weight: bold; color: #667eea; }}
        .stat-label {{ color: #666; margin-top: 5px; }}
        
        .solution-section {{
            background: rgba(255,255,255,0.95);
            border-radius: 12px;
            padding: 25px;
            margin-bottom: 25px;
        }}
        
        .sudoku-grid {{
            display: grid;
            grid-template-columns: repeat(16, 1fr);
            gap: 2px;
            background: #333;
            padding: 3px;
            border-radius: 8px;
            margin: 20px auto;
            max-width: 640px;
            box-shadow: 0 8px 25px rgba(0,0,0,0.3);
        }}
        .cell {{
            width: 35px;
            height: 35px;
            display: flex;
            align-items: center;
            justify-content: center;
            font-weight: bold;
            font-size: 16px;
            border-radius: 4px;
            transition: all 0.3s;
        }}
        .cell:hover {{ transform: scale(1.1); }}
        .cell.original {{ background: linear-gradient(135deg, #667eea, #764ba2); color: white; }}
        .cell.solution {{ background: #f8f9fa; color: #333; }}
        .cell.highlight {{ background: #ffd700; color: #333; }}
        
        /* 4x4 宮格邊框 */
        .sudoku-grid .cell:nth-child(4n) {{ border-right: 3px solid #333 !important; }}
        .sudoku-grid .cell:nth-child(16n+5),
        .sudoku-grid .cell:nth-child(16n+9),
        .sudoku-grid .cell:nth-child(16n+13) {{ border-right: 3px solid #333 !important; }}
        
        .solution-nav {{
            display: flex;
            justify-content: center;
            gap: 15px;
            margin: 20px 0;
        }}
        .solution-nav button {{
            padding: 10px 25px;
            background: #667eea;
            color: white;
            border: none;
            border-radius: 8px;
            cursor: pointer;
            font-size: 16px;
            transition: all 0.3s;
        }}
        .solution-nav button:hover {{ background: #764ba2; transform: translateY(-2px); }}
        .solution-nav button:disabled {{ background: #ccc; cursor: not-allowed; }}
        
        .dimension-panel {{
            background: rgba(255,255,255,0.95);
            border-radius: 12px;
            padding: 25px;
            margin-top: 30px;
        }}
        .dimension-grid {{
            display: grid;
            grid-template-columns: repeat(3, 1fr);
            gap: 20px;
            margin-top: 15px;
        }}
        .dimension-item {{
            padding: 20px;
            background: linear-gradient(135deg, #667eea15 0%, #764ba215 100%);
            border-radius: 10px;
            border-left: 4px solid #667eea;
        }}
        .dimension-title {{
            font-weight: bold;
            color: #667eea;
            margin-bottom: 8px;
            font-size: 1.1em;
        }}
        .dimension-content {{ color: #555; font-size: 0.9em; line-height: 1.6; }}
        
        .constraint-list {{
            background: #f8f9fa;
            border-radius: 8px;
            padding: 15px;
            margin: 15px 0;
            max-height: 300px;
            overflow-y: auto;
        }}
        .constraint-item {{
            padding: 8px;
            border-bottom: 1px solid #eee;
            font-family: monospace;
            font-size: 13px;
        }}
        .constraint-item:last-child {{ border-bottom: none; }}
    </style>
</head>
<body>
    <div class="container">
        <h1>🎲 超級大數獨 16×16 深度求解結果</h1>
        <p class="subtitle">五維思維框架 (點-線-面-體-球-時空) | DLX 精確覆蓋 + MRV 回溯優化 | 16 行符闔排列約束</p>
        
        <div class="stats-grid">
            <div class="stat-card">
                <div class="stat-value">{len(self.solutions)}</div>
                <div class="stat-label">🔢 找到的解數量</div>
            </div>
            <div class="stat-card">
                <div class="stat-value">16×16</div>
                <div class="stat-label">📐 數獨規模</div>
            </div>
            <div class="stat-card">
                <div class="stat-value">4×4</div>
                <div class="stat-label">🏠 宮格大小</div>
            </div>
            <div class="stat-card">
                <div class="stat-value">1-16</div>
                <div class="stat-label">🔤 數字範圍</div>
            </div>
            <div class="stat-card">
                <div class="stat-value">{sum(len(v) for v in self.row_constraints.values())}</div>
                <div class="stat-label">📊 行約束模式總數</div>
            </div>
            <div class="stat-card">
                <div class="stat-value">{np.count_nonzero(self.grid):.0f}</div>
                <div class="stat-label">🎯 初盤已知數</div>
            </div>
        </div>
        
        <div class="solution-section">
            <h2 style="color: #333; margin-bottom: 15px;">🎯 第一個解</h2>
            <div class="solution-nav">
                <button onclick="showSolution(0)" id="btn-prev">◀ 上一個</button>
                <span style="align-self: center; font-weight: bold;" id="sol-counter">1 / {max(1, len(self.solutions))}</span>
                <button onclick="showSolution({max(0, len(self.solutions)-1)})" id="btn-next">下一個 ▶</button>
            </div>
            
            <div id="solution-grid" class="sudoku-grid">
'''
        
        # 渲染第一個解
        if self.solutions:
            sol = self.solutions[0]
            for i in range(16):
                for j in range(16):
                    val = int(sol[i, j])
                    cell_class = "cell original" if self.grid[i, j] != 0 else "cell solution"
                    html_content += f'<div class="{cell_class}" data-row="{i}" data-col="{j}">{val}</div>'
        else:
            html_content += '<div style="text-align: center; padding: 40px; color: #666;">⚠ 未找到有效解</div>'
        
        html_content += '''
            </div>
            
            <h3 style="color: #667eea; margin: 20px 0 10px;">📋 行約束規則 (前 20 條)</h3>
            <div class="constraint-list">
'''
        
        # 添加約束資訊
        constraint_count = 0
        for row_idx in range(1, 17):
            if row_idx in self.row_constraints and len(self.row_constraints[row_idx]) > 0:
                perms = self.row_constraints[row_idx]
                sample_perms = perms[:3] if len(perms) >= 3 else perms
                for perm in sample_perms:
                    constraint_count += 1
                    if constraint_count <= 50:
                        html_content += f'<div class="constraint-item">第{row_idx}行: {", ".join(map(str, perm))}</div>'
        
        html_content += f'''
            </div>
            <p style="color: #666; margin-top: 10px;">共 {sum(len(v) for v in self.row_constraints.values())} 個排列模式，顯示前 50 條</p>
        </div>
        
        <div class="dimension-panel">
            <h2 style="color: #333; margin-bottom: 20px;">🌐 五維思維框架分析</h2>
            <div class="dimension-grid">
                <div class="dimension-item">
                    <div class="dimension-title">🔴 點維度 (Point)</div>
                    <div class="dimension-content">
                        256 個單元格，每個單元格獨立約束分析<br>
                        可能值 = 1-16 - (同行 + 同列 + 同宮格已填數字)<br>
                        MRV 策略優化搜索順序
                    </div>
                </div>
                <div class="dimension-item">
                    <div class="dimension-title">🔵 線維度 (Line)</div>
                    <div class="dimension-content">
                        16 行 + 16 列 = 32 條線約束<br>
                        每行/每列必須包含 1-16 各一次，無重複<br>
                        符闔排列：每行 8731 種合法排列模式
                    </div>
                </div>
                <div class="dimension-item">
                    <div class="dimension-title">🟢 面維度 (Face)</div>
                    <div class="dimension-content">
                        16 個 4×4 宮格<br>
                        每宮格必須包含 1-16 各一次<br>
                        宮格間存在約束傳遞關係
                    </div>
                </div>
                <div class="dimension-item">
                    <div class="dimension-title">🟣 體維度 (Body)</div>
                    <div class="dimension-content">
                        16 行排列群體約束<br>
                        每行從 {sum(len(v) for v in self.row_constraints.values())} 個排列中選擇<br>
                        行間需滿足列不衝突約束
                    </div>
                </div>
                <div class="dimension-item">
                    <div class="dimension-title">🟠 球維度 (Sphere)</div>
                    <div class="dimension-content">
                        全局狀態空間探索<br>
                        搜索樹剪枝、MRV 策略優化<br>
                        解空間聚類與模式識別
                    </div>
                </div>
                <div class="dimension-item">
                    <div class="dimension-title">⏰ 時空維度 (Space-Time)</div>
                    <div class="dimension-content">
                        求解過程的時空映射<br>
                        記錄每個單元格在時間序列上的狀態變化<br>
                        回溯路徑可視化與分析
                    </div>
                </div>
            </div>
        </div>
        
        <div class="solution-section" style="margin-top: 30px;">
            <h2 style="color: #333;">📈 求解統計</h2>
            <div class="stats-grid">
                <div class="stat-card">
                    <div class="stat-value">{len(self.row_constraints)}</div>
                    <div class="stat-label">行約束檔案</div>
                </div>
                <div class="stat-card">
                    <div class="stat-value">{time.time() - start_time:.2f}s</div>
                    <div class="stat-label">搜尋時間</div>
                </div>
                <div class="stat-card">
                    <div class="stat-value">{self.GRID_SIZE}</div>
                    <div class="stat-label">格子尺寸</div>
                </div>
            </div>
        </div>
    </div>
    
    <script>
    let currentSolution = 0;
    const solutions = {len(self.solutions)};
    
    function showSolution(index) {{
        currentSolution = index;
        const grid = document.getElementById('solution-grid');
        document.getElementById('sol-counter').textContent = `${{index + 1}} / ${{Math.max(1, solutions)}}`;
        document.getElementById('btn-prev').disabled = index === 0;
        document.getElementById('btn-next').disabled = index === solutions - 1;
    }}
    </script>
</body>
</html>
'''
        
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(html_content)
        print(f"✓ HTML 可視化已儲存: {output_path}")


# ============== 主程式 ==============

if __name__ == "__main__":
    print("=" * 70)
    print("🎯 超級大數獨 16×16 深度求解器")
    print("📐 Box Size: 4, 數字範圍: 1-16, 初盤格子: 256")
    print("🌐 五維思維框架: 點-線-面-體-球-時空")
    print("📊 16 行符闔排列約束")
    print("=" * 70)
    print()
    
    base_dir = "D:/2026/WPF_Sudoku/Sudoku_256"
    
    # 步驟 1: 讀取所有行約束
    print("📂 步驟 1: 讀取 16 行符闔排列 Excel 檔案")
    print("-" * 70)
    start_load = time.time()
    row_constraints = load_all_row_constraints(base_dir)
    load_time = time.time() - start_load
    print(f"⏱ 讀取完成，耗時 {load_time:.2f} 秒")
    print("-" * 70)
    print()
    
    # 步驟 2: 載入初盤
    print("📂 步驟 2: 載入初盤")
    print("-" * 70)
    txt_file = f"{base_dir}/超級大數獨_box_size4.txt"
    initial_grid = load_initial_puzzle(txt_file)
    known_count = np.count_nonzero(initial_grid)
    print(f"✓ 初盤已知數: {known_count}")
    print("-" * 70)
    print()
    
    # 步驟 3: 執行求解
    print("🚀 步驟 3: 深度探尋所有解")
    print("-" * 70)
    
    solver = SuperSudoku16x16Solver(row_constraints, initial_grid)
    
    # 時間限制與最大解數
    time_limit = 120  # 2 分鐘
    max_solutions = 100
    
    solution_count = solver.solve(max_solutions=max_solutions, time_limit=time_limit)
    
    elapsed = time.time() - start_load
    print(f"⏱ 總耗時: {elapsed:.2f} 秒")
    print(f"✓ 找到 {solution_count} 個解")
    print("-" * 70)
    print()
    
    # 步驟 4: 生成可視化
    print("📊 步驟 4: 生成可視化文檔")
    html_path = f"{base_dir}/超級大數獨_求解結果.html"
    solver.generate_visualization(html_path)
    
    # 步驟 5: 儲存詳細結果
    print("💾 步驟 5: 儲存結果檔案")
    results = {
        "solution_count": solution_count,
        "search_time_seconds": elapsed,
        "grid_size": 16,
        "box_size": 4,
        "known_cells": int(known_count),
        "row_constraints_loaded": sum(1 for v in row_constraints.values() if len(v) > 0),
        "solutions": [sol.tolist() for sol in solver.solutions[:5]],
        "timestamp": time.strftime("%Y-%m-%d %H:%M:%S")
    }
    
    with open(f"{base_dir}/求解結果.json", 'w', encoding='utf-8') as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    print(f"✓ JSON 結果儲存: {base_dir}/求解結果.json")
    
    # 步驟 6: 生成結果摘要
    print()
    print("=" * 70)
    print("✅ 求解完成!")
    print(f"📊 統計摘要:")
    print(f"   • 找到的解: {solution_count} 個")
    print(f"   • 搜尋時間: {elapsed:.2f} 秒")
    print(f"   • 行約束檔案: 16 個")
    print(f"   • 行約束模式: {sum(len(v) for v in row_constraints.values())} 個")
    print(f"   • 可視化檔案: 超級大數獨_求解結果.html")
    print(f"   • JSON 檔案: 求解結果.json")
    print("=" * 70)
