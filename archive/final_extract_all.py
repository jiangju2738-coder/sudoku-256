#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
最終修復版：正確提取 A6, A12, A13, A16 的排列資料
根據分析，這些檔案的排列分布在不同欄位中
"""

from openpyxl import load_workbook
import numpy as np
from collections import Counter

base_dir = "D:/2026/WPF_Sudoku/Sudoku_256"

# 檔案特定結構
FILE_CONFIGS = {
    6: {
        "name": "A6第六行符闔排列.xlsx",
        "data_pattern": "col4-17 + col19",  # 欄位 4-17 有 14 個數值 + 欄位 19 有 1 個數值 = 15 個，需要補齊 1 個
    },
    12: {
        "name": "A12第十二行符闔排列.xlsx",
        "data_pattern": "col4-19",  # 欄位 4-19 有 16 個數值
    },
    13: {
        "name": "A13第十三行符闔排列.xlsx",
        "data_pattern": "col4-19",
    },
    16: {
        "name": "A16第十六行符闔排列.xlsx",
        "data_pattern": "col4-19",
    },
}

def collect_row_permutation(row, row_idx):
    """收集一行的排列資料"""
    
    if row_idx == 6:
        # A6: 欄位 4-17 (14 個數值) + 欄位 19 (1 個數值) = 15 個數值
        values = []
        # 欄位 4-17
        for col in range(4, 18):
            if col < len(row):
                val = row[col]
                if isinstance(val, (int, float)) and 1 <= val <= 16:
                    values.append(int(val))
                elif isinstance(val, str) and val.startswith('='):
                    break
        
        # 欄位 19
        if 19 < len(row):
            val = row[19]
            if isinstance(val, (int, float)) and 1 <= val <= 16:
                values.append(int(val))
        
        # 如果只有 15 個數值，補齊缺失的數字
        if len(values) == 15:
            present = set(values)
            missing = [x for x in range(1, 17) if x not in present]
            if len(missing) == 1:
                values.append(missing[0])
        
        return values if len(values) == 16 else None
    
    elif row_idx in [12, 13, 16]:
        # 欄位 4-19 直接提取
        values = []
        for col in range(4, 20):
            if col < len(row):
                val = row[col]
                if isinstance(val, (int, float)) and 1 <= val <= 16:
                    values.append(int(val))
                elif isinstance(val, str) and val.startswith('='):
                    # 遇到公式，嘗試用欄位 18 的 ArrayFormula 值
                    # 或者查看其他欄位
                    if col == 18 and 19 < len(row):
                        val19 = row[19]
                        if isinstance(val19, (int, float)) and 1 <= val19 <= 16:
                            values.append(int(val19))
                    break
        
        return values if len(values) == 16 else None
    
    return None


def load_all_remaining(base_dir: str) -> dict:
    """載入所有剩餘檔案"""
    constraints = {}
    
    for row_idx in [6, 12, 13, 16]:
        config = FILE_CONFIGS[row_idx]
        filename = config["name"]
        filepath = f"{base_dir}/{filename}"
        
        print(f"\n{'='*70}")
        print(f"📊 最終讀取第{row_idx}行: {filename}")
        print(f"{'='*70}")
        
        try:
            wb = load_workbook(filepath, read_only=True)
            ws = wb.active
            
            print(f"📐 工作表: {ws.title}")
            
            # 先分析結構
            sample_row = None
            for row in ws.iter_rows(values_only=True):
                sample_row = row
                break
            
            if sample_row:
                print(f"📏 欄位數: {len(sample_row)}")
                print(f"🔍 欄位 4-19 的值:")
                for col in range(4, 20):
                    if col < len(sample_row):
                        val = sample_row[col]
                        if isinstance(val, (int, float)):
                            print(f"   欄位{col:2d}: {int(val)}", end="")
                            if 1 <= val <= 16:
                                print(" ✓", end="")
                            print()
                        else:
                            print(f"   欄位{col:2d}: {str(val)[:20]}", end="")
                            if isinstance(val, str) and val.startswith('='):
                                print(" [公式]", end="")
                            print()
                    else:
                        print(f"   欄位{col:2d}: N/A")
            
            wb.close()
            
            # 正式讀取
            wb = load_workbook(filepath, read_only=True)
            ws = wb.active
            
            row_permutations = []
            total_rows = 0
            
            for row in ws.iter_rows(values_only=True):
                total_rows += 1
                perm = collect_row_permutation(row, row_idx)
                if perm:
                    row_permutations.append(perm)
                
                if total_rows > 50000:
                    break
            
            wb.close()
            
            if row_permutations:
                constraints[row_idx] = np.array(row_permutations)
                print(f"\n✓ 成功讀取 {len(row_permutations):,} 個排列模式")
                for i, perm in enumerate(row_permutations[:3]):
                    print(f"  範例{i+1}: {perm}")
                    
                    # 驗證是否為完整排列
                    present = set(perm)
                    missing = [x for x in range(1, 17) if x not in present]
                    duplicates = [x for x in present if perm.count(x) > 1]
                    if missing or duplicates:
                        print(f"    ⚠ 驗證: 缺失{missing}, 重複{duplicates}")
                    else:
                        print(f"    ✓ 完整排列驗證通過")
            else:
                print(f"\n⚠ 未找到有效排列")
                
        except Exception as e:
            print(f"\n✗ 錯誤: {type(e).__name__}: {e}")
            import traceback
            traceback.print_exc()
            constraints[row_idx] = np.array([]).reshape(0, 16)
    
    return constraints


if __name__ == "__main__":
    print("=" * 70)
    print("🏁 最終修復版：提取剩餘 4 個檔案的排列資料")
    print("=" * 70)
    
    constraints = load_all_remaining(base_dir)
    
    print("\n" + "=" * 70)
    print("📊 最終彙總")
    print("=" * 70)
    
    total_perms = 0
    for row_idx in sorted(constraints.keys()):
        count = len(constraints[row_idx])
        total_perms += count
        status = "✓" if count > 0 else "⚠"
        print(f"   {status} 第{row_idx:2d}行: {count:>8,} 個排列")
    
    print(f"\n📈 新增排列模式總數: {total_perms:,}")
    
    # 儲存結果
    import json
    results = {str(k): len(v) for k, v in constraints.items() if len(v) > 0}
    with open(f"{base_dir}/最終提取結果.json", 'w') as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    print(f"✓ 結果已儲存到 最終提取結果.json")
