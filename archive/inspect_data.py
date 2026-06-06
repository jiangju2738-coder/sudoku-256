#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
詳細分析 Excel 檔案結構 - 逐行分析
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook

base_dir = "D:/2026/WPF_Sudoku/Sudoku_256"

# 只分析 A1 檔案來理解結構
filename = "A1第一行符闔排列.xlsx"
print(f"📂 分析 {filename}")

wb = load_workbook(f"{base_dir}/{filename}", read_only=True)
ws = wb.active

# 讀取所有資料到記憶體
print("📥 讀取所有資料...")
all_data = []
for row in ws.iter_rows(values_only=True):
    all_data.append(row)
wb.close()

print(f"📊 總共 {len(all_data)} 行, 每行 {len(all_data[0])} 列")

# 分析前 20 行資料
print("\n📋 前 20 行詳細資料:")
for i in range(min(20, len(all_data))):
    row = all_data[i]
    # 分析每列的類型
    row_analysis = []
    for j, val in enumerate(row[:20]):
        if val is None:
            row_analysis.append('NULL')
        elif isinstance(val, (int, float)):
            row_analysis.append(f'{val:.0f}')
        elif isinstance(val, str):
            row_analysis.append(f'"{val[:10]}"')
        else:
            row_analysis.append(str(val)[:10])
    print(f"  行{i:3d}: [{', '.join(row_analysis)}]")

# 分析欄位的含義
print("\n🔍 分析欄位結構:")
print(f"  欄位 0: 固定值 = {all_data[0][0]} (可能是行號)")
print(f"  欄位 1: 範圍 {min(r[1] for r in all_data)}-{max(r[1] for r in all_data)}")

# 欄位 2 可能包含字串，過濾後分析
col2_vals = [r[2] for r in all_data if isinstance(r[2], (int, float))]
if col2_vals:
    print(f"  欄位 2 (數值): 範圍 {min(col2_vals)}-{max(col2_vals)}")

col2_str = [r[2] for r in all_data if isinstance(r[2], str)]
if col2_str:
    print(f"  欄位 2 (字串): {set(col2_str[:10])}")

print(f"  欄位 3: 公式 = {all_data[0][3]}")

# 分析從欄位 4 開始的數值
print("\n📊 欄位 4-19 的數值分析:")
for col_idx in range(4, min(20, len(all_data[0]))):
    values = []
    for row in all_data[:100]:  # 取前 100 行樣本
        val = row[col_idx]
        if isinstance(val, (int, float)):
            values.append(val)
    
    if values:
        unique_vals = set(values)
        print(f"  欄位{col_idx}: {len(unique_vals)} 個唯一值, 範圍 {min(values)}-{max(values)}, 樣本 {list(unique_vals)[:5]}")

# 檢查是否有公式欄位
print("\n🔢 公式欄位分析:")
for col_idx in range(len(all_data[0])):
    sample_vals = [all_data[i][col_idx] for i in range(min(5, len(all_data)))]
    has_formula = any(isinstance(v, str) and v.startswith('=') for v in sample_vals if v is not None)
    if has_formula:
        print(f"  欄位{col_idx}: 包含公式 {sample_vals}")
