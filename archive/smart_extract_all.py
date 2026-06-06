#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
智能提取剩餘 4 個檔案的排列資料
A6, A12, A13, A16 - 這些檔案的 16 個數值分散在不同欄位
"""

from openpyxl import load_workbook
import numpy as np
from collections import Counter

base_dir = "D:/2026/WPF_Sudoku/Sudoku_256"
remaining_files = {
    6: "A6第六行符闔排列.xlsx",
    12: "A12第十二行符闔排列.xlsx",
    13: "A13第十三行符闔排列.xlsx",
    16: "A16第十六行符闔排列.xlsx"
}

# 基於深度分析發現的欄位模式
# 欄位結構：第0欄=行號, 第1欄=排列序號, 第2欄=重複, 第3欄=公式
# 真正的排列數值從第4欄開始，但有些被跳過了

def extract_permutation_smart(row, row_idx):
    """智能提取排列 - 尋找連續的 16 個 1-16 數值"""
    
    # 欄位 0, 1, 2, 3 是元資料，從欄位 4 開始尋找資料
    # 但有些檔案在第 4-5 欄有重複的排列標識
    
    # 方法：找出所有 1-16 的數值，驗證是否構成完整排列
    all_numeric = []
    for val in row:
        if isinstance(val, (int, float)) and 1 <= val <= 16:
            all_numeric.append(int(val))
    
    # 如果數量 >= 16，嘗試找出 16 個唯一值的排列
    if len(all_numeric) >= 16:
        # 統計每個數字出現次數
        counts = Counter(all_numeric)
        unique_values = set(counts.keys())
        
        # 如果剛好包含 1-16 各至少一次，構建排列
        if len(unique_values) >= 16:
            # 從欄位 4 開始，尋找連續的排列
            data_start = 4
            data_vals = []
            
            for col_idx in range(data_start, min(len(row), data_start + 20)):
                val = row[col_idx]
                if isinstance(val, (int, float)) and 1 <= val <= 16:
                    data_vals.append(int(val))
                elif isinstance(val, str) and val.startswith('='):
                    # 遇到公式就停止
                    break
            
            if len(data_vals) >= 16:
                # 檢查是否包含 1-16 的完整排列
                present = set(data_vals[:16])
                if present == set(range(1, 17)):
                    return data_vals[:16]
        
        # 如果欄位 4 開始的資料不符合，嘗試其他位置
        # 查找是否有連續的 16 個不同數值
        for start in range(len(all_numeric) - 15):
            candidate = all_numeric[start:start+16]
            if len(set(candidate)) == 16 and set(candidate) == set(range(1, 17)):
                return candidate
    
    return None


def load_remaining_constraints(base_dir: str) -> dict:
    """載入剩餘 4 個檔案"""
    constraints = {}
    
    for row_idx, filename in remaining_files.items():
        filepath = f"{base_dir}/{filename}"
        
        print(f"\n{'='*70}")
        print(f"📊 智能讀取第{row_idx}行: {filename}")
        print(f"{'='*70}")
        
        try:
            wb = load_workbook(filepath, read_only=True)
            ws = wb.active
            
            print(f"📐 工作表: {ws.title}, 最大欄位: {ws.max_column}")
            
            # 先分析前 5 行的資料結構
            sample_rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                sample_rows.append(row)
                if i >= 5:
                    break
            
            # 分析每行的 1-16 數值位置
            print("\n🔍 資料結構分析:")
            for i, row in enumerate(sample_rows):
                data_positions = []
                for j, val in enumerate(row):
                    if isinstance(val, (int, float)) and 1 <= val <= 16:
                        data_positions.append(j)
                print(f"  行{i}: 欄位數={len(row)}, 1-16數值在欄位: {data_positions[:10]}...")
            
            wb.close()
            
            # 根據分析結果，重新載入
            wb = load_workbook(filepath, read_only=True)
            ws = wb.active
            
            row_permutations = []
            total_rows = 0
            max_rows = 30000  # 限制讀取行數
            
            # 重新分析檔案結構
            file_structure = None
            for row in ws.iter_rows(values_only=True):
                total_rows += 1
                if total_rows <= 5:
                    # 記錄結構
                    data_positions = []
                    for j, val in enumerate(row):
                        if isinstance(val, (int, float)) and 1 <= val <= 16:
                            data_positions.append(j)
                    if file_structure is None:
                        file_structure = data_positions
                else:
                    break
            
            print(f"\n📋 檔案結構: 1-16數值在欄位 {file_structure}")
            
            # 重新讀取所有行
            wb.close()
            wb = load_workbook(filepath, read_only=True)
            ws = wb.active
            
            if file_structure and len(file_structure) >= 16:
                for row in ws.iter_rows(values_only=True):
                    if len(row) > max(file_structure):
                        # 從記錄的位置提取數值
                        perm = [int(row[pos]) for pos in file_structure[:16] 
                                if pos < len(row) and isinstance(row[pos], (int, float))]
                        if len(perm) == 16 and set(perm) == set(range(1, 17)):
                            row_permutations.append(perm)
                    else:
                        # 智能提取
                        perm = extract_permutation_smart(row, row_idx)
                        if perm:
                            row_permutations.append(perm)
                    
                    if total_rows > 50000:
                        break
            
            wb.close()
            
            if row_permutations:
                constraints[row_idx] = np.array(row_permutations)
                print(f"✓ 成功讀取 {len(row_permutations):,} 個排列模式")
                for i, perm in enumerate(row_permutations[:3]):
                    print(f"  範例{i+1}: {perm}")
            else:
                print(f"⚠ 未找到有效排列")
                
        except Exception as e:
            print(f"✗ 錯誤: {type(e).__name__}: {e}")
            constraints[row_idx] = np.array([]).reshape(0, 16)
    
    return constraints


if __name__ == "__main__":
    print("=" * 70)
    print("🧠 智能提取剩餘 4 個檔案的排列資料")
    print("=" * 70)
    
    constraints = load_remaining_constraints(base_dir)
    
    print("\n" + "=" * 70)
    print("📊 最終結果彙總")
    print("=" * 70)
    
    total_perms = 0
    for row_idx in sorted(constraints.keys()):
        count = len(constraints[row_idx])
        total_perms += count
        status = "✓" if count > 0 else "⚠"
        print(f"   {status} 第{row_idx:2d}行: {count:>8,} 個排列")
    
    print(f"\n📈 新增排列模式總數: {total_perms:,}")
    
    # 儲存結果
    import json
    results = {str(k): len(v) for k, v in constraints.items() if len(v) > 0}
    with open(f"{base_dir}/智能提取結果.json", 'w') as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    print(f"✓ 結果已儲存到 智能提取結果.json")
