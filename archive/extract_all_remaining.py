#!/usr/bin/env python3
"""
高效提取所有剩余行的约束排列
处理A1, A2, A4, A5, A6-A16 (除A3外所有行)
"""

import json
import os
import openpyxl
from collections import defaultdict

BASE = "D:/2026/WPF_Sudoku/Sudoku_256/"

def extract_permutations(filepath, row_num):
    """提取单行16列排列，处理ArrayFormula情况"""
    permutations = []
    
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active
    
    for row in ws.iter_rows(min_row=1, min_col=5, max_col=20, values_only=True):
        # E-T列 (索引4-19) = 16个数字
        values = list(row)
        numeric_values = []
        
        for v in values:
            if v is None:
                continue
            if isinstance(v, (int, float)) and 1 <= v <= 16:
                numeric_values.append(int(v))
            elif isinstance(v, str) and v.startswith('='):
                # 跳过公式
                continue
        
        if len(numeric_values) == 16:
            # 完整16个数字
            if all(1 <= x <= 16 for x in numeric_values):
                permutations.append(tuple(numeric_values))
        elif len(numeric_values) == 15:
            # 15个数字 + 1个ArrayFormula, 补上缺失的数字
            present = set(numeric_values)
            missing = [x for x in range(1, 17) if x not in present]
            if len(missing) == 1:
                perm = list(numeric_values)
                perm.append(missing[0])
                permutations.append(tuple(perm))
    
    wb.close()
    return permutations


def extract_in_chunks(filepath, chunk_size=50000):
    """分块提取，内存友好"""
    all_perms = []
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active
    
    for row in ws.iter_rows(min_row=1, min_col=5, max_col=20, values_only=True):
        values = list(row)
        numeric_values = []
        
        for v in values:
            if v is None:
                continue
            if isinstance(v, (int, float)) and 1 <= v <= 16:
                numeric_values.append(int(v))
            elif isinstance(v, str) and v.startswith('='):
                continue
        
        if len(numeric_values) == 16:
            if all(1 <= x <= 16 for x in numeric_values):
                all_perms.append(tuple(numeric_values))
        elif len(numeric_values) == 15:
            present = set(numeric_values)
            missing = [x for x in range(1, 17) if x not in present]
            if len(missing) == 1:
                perm = list(numeric_values)
                perm.append(missing[0])
                all_perms.append(tuple(perm))
        
        # 每chunk_size行打印进度
        if len(all_perms) % chunk_size == 0 and len(all_perms) > 0:
            print(f"  已处理 {len(all_perms):,} 个排列...")
    
    wb.close()
    return all_perms


def find_excel_file(row_num):
    """找到对应的Excel文件"""
    for f in os.listdir(BASE):
        if f.startswith(f"A{row_num}") and f.endswith(".xlsx"):
            return os.path.join(BASE, f)
    return None


def main():
    # 需要处理的行 (跳过A3, 已处理)
    rows_to_process = [1, 2, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16]
    
    # 存储所有结果
    all_results = defaultdict(list)
    progress = {}
    
    for row_num in rows_to_process:
        filepath = find_excel_file(row_num)
        if not filepath:
            print(f"✗ A{row_num}: 未找到Excel文件")
            continue
        
        file_size_mb = os.path.getsize(filepath) / 1024 / 1024
        print(f"\n{'='*60}")
        print(f"处理 A{row_num} ({file_size_mb:.1f} MB)...")
        
        try:
            if file_size_mb > 100:
                # 大文件使用分块处理
                perms = extract_in_chunks(filepath)
            else:
                # 小文件直接提取
                perms = extract_permutations(filepath, row_num)
            
            all_results[f"A{row_num}"] = perms
            progress[f"A{row_num}"] = {
                "count": len(perms),
                "status": "completed"
            }
            print(f"✓ A{row_num}: {len(perms):,} 个排列")
            
            # 保存单个JSON文件
            output_path = os.path.join(BASE, f"A{row_num}_permutations.json")
            with open(output_path, 'w') as f:
                json.dump(perms, f)
            print(f"  已保存: A{row_num}_permutations.json")
            
        except Exception as e:
            print(f"✗ A{row_num}: 错误 - {e}")
            progress[f"A{row_num}"] = {
                "count": 0,
                "status": "error",
                "error": str(e)
            }
    
    # 保存进度文件
    progress_path = os.path.join(BASE, "完整16行約束_進度.json")
    
    # 先读取现有的进度数据 (除了A3)
    existing_data = {}
    if os.path.exists(progress_path):
        try:
            # 使用流式读取来避免内存问题
            with open(progress_path, 'r') as f:
                content = f.read()
                # 移除A3的数据，合并新的数据
                # 由于A3已经单独保存，我们更新progress部分
                existing_data = json.loads(content.split('"row_constraints"')[0] + '"row_constraints": {} }')
        except:
            pass
    
    # 创建完整进度
    final_progress = {
        "row_constraints": dict(all_results),
        "progress": "A3已单独保存，其他行已更新",
        "total_patterns": sum(len(v) for v in all_results.values()) + 407669,
        "per_row": {k: len(v) for k, v in all_results.items()},
        "A3_status": "已单独保存在A3_permutations.json (407,669个排列)"
    }
    
    # 保存完整进度 (注意: 这会创建大文件)
    # 为了节省空间，我们只保存进度元数据，数据已单独保存
    progress_meta = {
        "status": "completed",
        "A3": {"file": "A3_permutations.json", "count": 407669},
    }
    for k, v in progress.items():
        progress_meta[k] = v
    
    with open(os.path.join(BASE, "提取进度_更新.json"), 'w') as f:
        json.dump(progress_meta, f, indent=2)
    
    print(f"\n{'='*60}")
    print("提取完成!")
    print(f"总排列数: {final_progress['total_patterns']:,}")
    
    for row in sorted(progress.keys()):
        print(f"  {row}: {progress[row]['count']:,}")
    
    print(f"\n单文件JSON已保存:")
    for row in sorted(all_results.keys()):
        print(f"  A{row}_permutations.json")


if __name__ == "__main__":
    main()
