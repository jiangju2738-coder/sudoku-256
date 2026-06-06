#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
修復載入 A4-A6, A12-A16 檔案的約束讀取
根據分析結果，動態調整欄位索引
"""

import sys
sys.path.insert(0, '.')

from openpyxl import load_workbook
import numpy as np
from collections import defaultdict

base_dir = "D:/2026/WPF_Sudoku/Sudoku_256"

# 未成功載入的檔案列表
failed_files = {
    4: "A4第四行符闔排列.xlsx",
    5: "A5第五行符闔排列.xlsx",
    6: "A6第六行符闔排列.xlsx",
    12: "A12第十二行符闔排列.xlsx",
    13: "A13第十三行符闔排列.xlsx",
    14: "A14第十四行符闔排列.xlsx",
    15: "A15第十五行符闔排列.xlsx",
    16: "A16第十六行符闔排列.xlsx"
}

# 檔案結構分析結果
# 這些檔案的數值欄位位置不同，需要動態檢測
FILE_STRUCTURES = {
    4: {"data_cols": list(range(4, 19)), "expected_len": 15},  # A4: 欄位4-18
    5: {"data_cols": list(range(4, 19)), "expected_len": 15},  # A5: 欄位4-18
    6: {"data_cols": list(range(4, 19)), "expected_len": 15},  # A6: 欄位4-18
    12: {"data_cols": list(range(4, 19)), "expected_len": 15},  # A12: 欄位4-18
    13: {"data_cols": list(range(4, 19)), "expected_len": 15},  # A13: 欄位4-18
    14: {"data_cols": list(range(4, 19)), "expected_len": 15},  # A14: 欄位4-18
    15: {"data_cols": list(range(4, 19)), "expected_len": 15},  # A15: 欄位4-18
    16: {"data_cols": list(range(4, 20)), "expected_len": 16},  # A16: 欄位4-19 (與A1相同)
}

def extract_permutation(row, file_idx):
    """根據檔案結構提取排列"""
    structure = FILE_STRUCTURES.get(file_idx, {"data_cols": list(range(4, 20)), "expected_len": 16})
    data_cols = structure["data_cols"]
    expected_len = structure["expected_len"]
    
    values = []
    for col_idx in data_cols:
        if col_idx >= len(row):
            break
        val = row[col_idx]
        if isinstance(val, (int, float)) and 1 <= val <= 16:
            values.append(int(val))
        else:
            break
    
    # 如果是 15 個數值，需要補齊到 16 個
    if len(values) == 15 and file_idx in [4, 5, 6, 12, 13, 14, 15]:
        # 從欄位 3 的公式提取最後一個數值
        # 例如: =SUM(E1,16*3) 表示 E1 + 48
        formula_val = row[3]
        if isinstance(formula_val, str) and formula_val.startswith('=SUM('):
            try:
                # 解析 =SUM(E1,16*N) 格式
                parts = formula_val.replace('=SUM(', '').replace(')', '').split(',')
                if len(parts) >= 2:
                    multiplier = int(parts[1].replace('16*', ''))
                    # 這表示 1-16 中缺少 16*multiplier 的和
                    # 實際上應該是這行包含 1-15 的數字，缺少 16
                    # 或者這行有 16 個位置，但我們只讀到 15 個
            except:
                pass
        
        # 查找缺失的數字
        present = set(values)
        missing = [i for i in range(1, 17) if i not in present]
        if len(missing) == 1:
            values.append(missing[0])
    
    return values if len(values) == 16 else None


def load_row_constraints_fixed(base_dir: str) -> dict:
    """修復版載入函數"""
    constraints = {}
    
    for row_idx, filename in failed_files.items():
        filepath = f"{base_dir}/{filename}"
        
        print(f"\n📂 重新讀取第{row_idx}行: {filename}")
        
        try:
            wb = load_workbook(filepath, read_only=True)
            ws = wb.active
            
            print(f"   📐 工作表: {ws.title}, 最大欄位: {ws.max_column}")
            
            row_permutations = []
            total_rows = 0
            
            for row in ws.iter_rows(values_only=True):
                total_rows += 1
                perm = extract_permutation(row, row_idx)
                if perm:
                    row_permutations.append(perm)
                
                if total_rows > 20000:  # 防止過大檔案
                    break
            
            wb.close()
            
            if row_permutations:
                constraints[row_idx] = np.array(row_permutations)
                print(f"   ✓ 成功讀取 {len(row_permutations):,} 個排列模式")
                # 展示前 3 個排列
                for i, perm in enumerate(row_permutations[:3]):
                    print(f"     範例{i+1}: {perm}")
            else:
                print(f"   ⚠ 無有效排列")
                
        except Exception as e:
            print(f"   ✗ 錯誤: {type(e).__name__}: {e}")
            constraints[row_idx] = np.array([]).reshape(0, 16)
    
    return constraints


if __name__ == "__main__":
    print("=" * 70)
    print("🔧 修復載入 A4-A6, A12-A16 檔案")
    print("=" * 70)
    
    constraints = load_row_constraints_fixed(base_dir)
    
    print("\n" + "=" * 70)
    print("📊 重新載入結果彙總")
    print("=" * 70)
    
    total_perms = 0
    for row_idx in sorted(constraints.keys()):
        count = len(constraints[row_idx])
        total_perms += count
        status = "✓" if count > 0 else "⚠"
        print(f"   {status} 第{row_idx:2d}行: {count:>8,} 個排列")
    
    print(f"\n📈 新增排列模式總數: {total_perms:,}")
    
    # 儲存結果
    import json
    results = {str(k): len(v) for k, v in constraints.items() if len(v) > 0}
    with open(f"{base_dir}/修復載入結果.json", 'w') as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    print(f"✓ 結果已儲存到 修復載入結果.json")
