#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
深度探索 16 行符闔排列的約束數據
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
import json
from collections import Counter

base_dir = "D:/2026/WPF_Sudoku/Sudoku_256"

# 檔案列表
file_list = [
    "A1第一行符闔排列.xlsx",
    "A2第二行符闔排列.xlsx",
    "A3第三行符闔排列.xlsx",
    "A4第四行符闔排列.xlsx",
    "A5第五行符闔排列.xlsx",
    "A6第六行符闔排列.xlsx",
    "A7第七行符闔排列.xlsx",
    "A8第八行符闔排列.xlsx",
    "A9第九行符闔排列.xlsx",
    "A10第十行符闔排列.xlsx",
    "A11第十一行符闔排列.xlsx",
    "A12第十二行符闔排列.xlsx",
    "A13第十三行符闔排列.xlsx",
    "A14第十四行符闔排列.xlsx",
    "A15第十五行符闔排列.xlsx",
    "A16第十六行符闔排列.xlsx"
]

# 分析檔案結構
analysis_results = {}

for filename in file_list:
    row_idx = int(''.join(filter(str.isdigit, filename.split('第')[0])))
    
    print(f"\n{'='*60}")
    print(f"📊 分析 {filename} (第{row_idx}行)")
    print(f"{'='*60}")
    
    wb = load_workbook(f"{base_dir}/{filename}", read_only=True)
    ws = wb.active
    
    # 讀取所有資料
    all_data = []
    for row in ws.iter_rows(values_only=True):
        all_data.append(row)
    
    df = pd.DataFrame(all_data)
    wb.close()
    
    print(f"📏 尺寸: {df.shape[0]} 行 × {df.shape[1]} 列")
    print(f"📝 欄位數: {df.shape[1]}")
    
    # 前 10 行預覽
    print(f"\n📋 前 10 行預覽:")
    for i in range(min(10, len(all_data))):
        row = all_data[i]
        non_null = [str(v) for v in row[:16] if v is not None]
        print(f"  {i+1}: {non_null}")
    
    # 分析數值分佈
    numeric_cols = []
    for col in range(16, df.shape[1]):  # 從第 17 欄開始是數值欄位 (E-T)
        try:
            col_data = pd.to_numeric(df[col], errors='coerce')
            if not col_data.isna().all():
                numeric_cols.append(col)
                col_values = col_data.dropna().unique()
                print(f"  欄位{col} (E+{col-16}): {len(col_values)} 個唯一值, 範圍 {col_values.min():.0f}-{col_values.max():.0f}")
        except:
            pass
    
    analysis_results[f"row_{row_idx}"] = {
        "filename": filename,
        "shape": df.shape,
        "sample_rows": len(all_data)
    }
    
    # 儲存分析結果
    if row_idx == 1:
        sample_data = all_data[:100]
        with open(f"{base_dir}/sample_data_row1.json", 'w') as f:
            json.dump(sample_data, f, ensure_ascii=False)

print(f"\n{'='*60}")
print("✅ 分析完成")
print(f"{'='*60}")

# 儲存總體分析
with open(f"{base_dir}/constraint_analysis.json", 'w') as f:
    json.dump(analysis_results, f, ensure_ascii=False, indent=2)
print("📄 分析結果已儲存到 constraint_analysis.json")
